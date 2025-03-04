# crm/export_utils.py
import csv
from io import BytesIO
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
import datetime
from pytz import timezone as pytz_timezone
from django.utils import timezone



def exportar_clientes(queryset, formato):
    # Campos a exportar
    CAMPOS = [
        ('nombre', 'Nombre'),
        ('apellido', 'Apellido'),
        ('cedula_pasaporte', 'Cédula/Pasaporte'),
        ('fecha_nacimiento', 'Fecha Nacimiento'),
        ('nacionalidad', 'Nacionalidad'),
        ('lugar_trabajo', 'Lugar de Trabajo'),
        ('cargo', 'Cargo'),
        ('email', 'Email'),
        ('direccion_fisica', 'Dirección'),
        ('telefono', 'Teléfono'),
        ('movil', 'Móvil'),
        ('fecha_creacion', 'Fecha Creación'),
        ('ultima_actividad', 'Última Actividad'),
        ('estado', 'Estado')
    ]

    def get_row_data(cliente):
        return [
            cliente.nombre,
            cliente.apellido,
            cliente.cedula_pasaporte,
            cliente.fecha_nacimiento.strftime("%d/%m/%Y") if cliente.fecha_nacimiento else '',
            str(cliente.nacionalidad.name),  # Aseguramos string
            cliente.lugar_trabajo or '',
            cliente.cargo or '',
            cliente.email,
            cliente.direccion_fisica,
            cliente.telefono,
            cliente.movil,
            cliente.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            cliente.ultima_actividad.strftime("%d/%m/%Y %H:%M"),
            cliente.get_estado_display()
        ]

    if formato == 'csv':
        response = HttpResponse(
            content_type='text/csv; charset=utf-8-sig',  # BOM para Excel
            headers={'Content-Disposition': 'attachment; filename="clientes_export.csv"'},
        )
        
        writer = csv.writer(response, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        
        writer.writerow([header for _, header in CAMPOS])
        
        for cliente in queryset.iterator():
            writer.writerow(get_row_data(cliente))
        
        return response

    elif formato == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="clientes_export.xlsx"'},
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", 
            end_color="4F81BD", 
            fill_type="solid"
        )
        alignment = Alignment(
            horizontal="center", 
            vertical="center", 
            wrap_text=True
        )
        
        # Encabezados con estilo
        for col_num, (_, header) in enumerate(CAMPOS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        # Datos
        for row_num, cliente in enumerate(queryset.iterator(), 2):
            # Convertir fechas timezone-aware a naive
            fecha_creacion = cliente.fecha_creacion
            if fecha_creacion and timezone.is_aware(fecha_creacion):
                fecha_creacion = timezone.make_naive(
                    fecha_creacion, 
                    timezone.get_current_timezone()
                )
            
            ultima_actividad = cliente.ultima_actividad
            if ultima_actividad and timezone.is_aware(ultima_actividad):
                ultima_actividad = timezone.make_naive(
                    ultima_actividad,
                    timezone.get_current_timezone()
                )
            
            row_data = [
                cliente.nombre,
                cliente.apellido,
                cliente.cedula_pasaporte,
                cliente.fecha_nacimiento,  # Date object
                str(cliente.nacionalidad.name),
                cliente.lugar_trabajo or '',
                cliente.cargo or '',
                cliente.email,
                cliente.direccion_fisica,
                cliente.telefono,
                cliente.movil,
                fecha_creacion,          # DateTime object
                ultima_actividad,        # DateTime object
                cliente.get_estado_display()
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
    
    # Formatear fechas
                if col_num == 4 and isinstance(value, datetime.date):
                    cell.number_format = "dd/mm/yyyy"  # Formato personalizado para fecha con año en 4 dígitos
                elif col_num in [12, 13] and isinstance(value, datetime.datetime):
                    cell.number_format = "dd/mm/yyyy hh:mm"  # Formato personalizado para fecha y hora
        
        # Ajustar anchos de columnas
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar encabezado y agregar filtros
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
        buffer = BytesIO()
        wb.save(buffer)  # Guardar el workbook en el buffer
        
        # ¡Mover el puntero al inicio del buffer!
        buffer.seek(0)
        
        response.write(buffer.getvalue())
        return response

def exportar_empresas(queryset, formato):
    CAMPOS = [
        ('nombre_comercial', 'Nombre Comercial'),
        ('razon_social', 'Razón Social'),
        ('rnc', 'RNC'),
        ('direccion_fisica', 'Dirección Física'),
        ('direccion_electronica', 'Email'),
        ('telefono', 'Teléfono Principal'),
        ('telefono2', 'Teléfono Secundario'),
        ('sitio_web', 'Sitio Web'),
        ('representante', 'Representante'),
        ('cargo_representante', 'Cargo Representante'),
        ('fecha_registro', 'Fecha Registro'),
        ('ultima_actividad', 'Última Actividad'),
        ('estado', 'Estado')
    ]

    def get_row_data(empresa):
        return [
            empresa.nombre_comercial,
            empresa.razon_social,
            empresa.rnc,
            empresa.direccion_fisica,
            empresa.direccion_electronica,
            empresa.telefono,
            empresa.telefono2 or '',
            empresa.sitio_web or '',
            empresa.representante,
            empresa.cargo_representante,
            empresa.fecha_registro.strftime("%d/%m/%Y %H:%M"),
            empresa.ultima_actividad.strftime("%d/%m/%Y %H:%M"),
            empresa.get_estado_display()
        ]

    if formato == 'csv':
        response = HttpResponse(
            content_type='text/csv; charset=utf-8-sig',
            headers={'Content-Disposition': 'attachment; filename="empresas_export.csv"'},
        )
        
        writer = csv.writer(response, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        writer.writerow([header for _, header in CAMPOS])
        
        for empresa in queryset.iterator():
            writer.writerow(get_row_data(empresa))
        
        return response

    elif formato == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="empresas_export.xlsx"'},
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Encabezados
        for col_num, (_, header) in enumerate(CAMPOS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        # Datos
        # Dentro del loop de empresas en excel:
        for row_num, empresa in enumerate(queryset.iterator(), 2):
    # Convertir fechas a naive datetime CORREGIDO
            fecha_registro = empresa.fecha_registro
            if fecha_registro and timezone.is_aware(fecha_registro):
                fecha_registro = timezone.make_naive(
                    fecha_registro,
                    timezone.get_current_timezone()  # Agregar zona horaria
                )
    
            ultima_actividad = empresa.ultima_actividad
            if ultima_actividad and timezone.is_aware(ultima_actividad):
                ultima_actividad = timezone.make_naive(
                    ultima_actividad,
                    timezone.get_current_timezone()  # Agregar zona horaria
                )
    
            row_data = [
                empresa.nombre_comercial,
                empresa.razon_social,
                empresa.rnc,
                empresa.direccion_fisica,
                empresa.direccion_electronica,
                empresa.telefono,
                empresa.telefono2 or '',
                empresa.sitio_web or '',
                empresa.representante,
                empresa.cargo_representante,
                fecha_registro,  # Ahora es datetime naive
                ultima_actividad,  # datetime naive
                empresa.get_estado_display()
            ]
    
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
        
        # Formatear fechas
                if col_num in [11, 12] and isinstance(value, datetime.datetime):
                    cell.number_format = "dd/mm/yyyy hh:mm"
        
        # Ajustar anchos, congelar, filtros (igual que clientes)
                for col in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try: max_length = max(max_length, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response